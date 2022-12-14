import datetime

from openpyxl.reader.excel import load_workbook

from studying_process.models import DirectionOfTraining, Group


class ExcelFile:
    gender_map = {'man': 'м', 'women': 'ж'}

    def __init__(self):
        self.filename = f'{datetime.datetime.now()}'
        self.work_book = load_workbook('sample.xlsx')
        self.ws = self.work_book.active

    def __str__(self) -> str:
        return self.filename

    def add_groups(self) -> 'ExcelFile':
        """ Добавить информафию о группах в файл """
        groups = Group.objects.prefetch_related('students').all().order_by('title')
        index_str = 2
        for count_group, group in enumerate(groups, start=1):
            self.ws[f'E{index_str}'] = f'{count_group}.{group.title} ({len(group.students.all())}/{group.max_length})'
            self.ws[f'F{index_str}'] = group.max_length - len(group.students.all())
            self.ws[f'G{index_str}'] = len(group.students.filter(gender='man'))
            self.ws[f'H{index_str}'] = len(group.students.filter(gender='women'))
            index_str += 1
            for count_student, student in enumerate(group.students.all().order_by('surname', 'name'), start=1):
                self.ws[f'E{index_str}'] = f'   {count_student}.{student} ({self.gender_map[student.gender]})'
                index_str += 1
        return self

    def add_direction_of_training(self) -> 'ExcelFile':
        """ Добавить информафию о направлениях в файл """
        directions = DirectionOfTraining.objects.select_related('curator').prefetch_related(
            'disciplines').all().order_by('title')
        index_str = 2
        for count_direction, direction in enumerate(directions, start=1):
            self.ws[f'A{index_str}'] = f'{count_direction}.{direction.title}'
            self.ws[f'B{index_str}'] = direction.curator.get_full_name() or direction.curator.username
            index_str += 1
            for count_discipline, discipline in enumerate(direction.disciplines.all(), start=1):
                self.ws[f'A{index_str}'] = f'   {count_discipline}.{discipline.title}'
                index_str += 1
        return self

    def save_file(self):
        """ Сохранить файл """
        self.work_book.save(f"/app/{self.filename}.xlsx")
